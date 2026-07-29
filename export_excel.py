"""
export_excel.py — Генерира/обновява Excel файл с два таба:

  1. "Разговори"              — детайлен ред на всеки анализиран разговор,
                                 с AutoFilter (дата, служител, статус) и
                                 условно оцветяване (зелено/червено).
  2. "Служители - обобщение"  — по един ред на служител: брой разговори,
                                 среден резултат, % флагнати, среден
                                 резултат по всеки критерий. Формулите
                                 (AVERAGEIFS/COUNTIFS) сочат към таб 1 и
                                 се преизчисляват автоматично при промяна.

ВАЖНО: Файлът се ПРЕГЕНЕРИРА изцяло от базата данни при всяко пускане —
базата (calls.db) е източникът на истината, не Excel файлът. Ако някой
ръчно редактира Excel файла, промените ще се загубят при следващия export.
"""

import subprocess
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

import db
from criteria_config import CRITERIA

OUTPUT_PATH = Path(__file__).parent / "call_analysis_report.xlsx"

FONT_NAME = "Arial"
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
FLAGGED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GOOD_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

# Фиксирани колони в таб "Разговори", преди динамичните колони по критерий
DETAIL_FIXED_HEADERS = [
    "Дата", "Служител", "ID на разговор", "Продължителност (мин)",
    "Посока", "Спомената 'Корект'", "Категория", "Тип разговор", "Статус", "Общ резултат",
]
DETAIL_TAIL_HEADERS = ["Причина за флаг", "Обобщение"]

CATEGORY_LABELS = {
    "нова_поръчка": "Нова поръчка",
    "потвърждение": "Потвърждение",
    "нерелевантен": "Нерелевантен",
    "друго": "Друго",
}

CALL_TYPE_LABELS = {
    "бърза_оферта": "Бърза оферта",
    "пълна_консултация": "Пълна консултация",
}


def _style_header_row(ws, row_idx: int, n_cols: int):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def build_detail_sheet(ws, analyzed_calls):
    criterion_labels = [c["label"] for c in CRITERIA]
    headers = DETAIL_FIXED_HEADERS + criterion_labels + DETAIL_TAIL_HEADERS
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    n_fixed = len(DETAIL_FIXED_HEADERS)

    for call in analyzed_calls:
        scores = {}
        for s in db.get_criteria_scores_for_analysis(call["analysis_id"]):
            scores[s["criterion_key"]] = s["score"] if s["applicable"] else "Неприложимо"
        if call["call_category"] == "нерелевантен":
            status = "Нерелевантен"
        else:
            status = "Флагнат" if call["is_flagged"] else "Успешен"
        category_label = CATEGORY_LABELS.get(call["call_category"], call["call_category"] or "")
        call_type_label = CALL_TYPE_LABELS.get(call["call_type"], call["call_type"] or "")
        korekt_label = "Да" if call["korekt_mentioned"] else "Не"
        row = [
            call["call_date"],
            call["agent_name"] or "",
            call["external_call_id"],
            round((call["duration_seconds"] or 0) / 60, 1),
            call["direction"] or "",
            korekt_label,
            category_label,
            call_type_label,
            status,
            call["overall_score"],
        ]
        row += [scores.get(c["key"], "") for c in CRITERIA]
        row += [call["flag_reason"] or "", call["overall_summary"] or ""]
        ws.append(row)

        for col in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=col).font = Font(name=FONT_NAME)

    last_row = ws.max_row
    last_col = len(headers)
    last_col_letter = get_column_letter(last_col)

    ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"
    ws.freeze_panes = "A2"

    status_col_letter = get_column_letter(9)
    data_range = f"A2:{last_col_letter}{last_row}"
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(formula=[f'${status_col_letter}2="Флагнат"'], fill=FLAGGED_FILL),
    )
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(formula=[f'${status_col_letter}2="Успешен"'], fill=GOOD_FILL),
    )
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(formula=[f'${status_col_letter}2="Нерелевантен"'],
                    fill=PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")),
    )

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18
    for i, _ in enumerate(criterion_labels, start=n_fixed + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16
    ws.column_dimensions[get_column_letter(last_col - 1)].width = 30
    ws.column_dimensions[get_column_letter(last_col)].width = 40

    return last_row, last_col_letter


def build_summary_sheet(ws, agent_names, detail_last_row, detail_last_col_letter):
    """
    Таб 'Служители - обобщение'. Формулите сочат към таб 'Разговори'
    (AVERAGEIFS/COUNTIFS) — преизчисляват се автоматично при нови данни.
    """
    headers = ["Служител", "Брой разговори", "Среден резултат",
               "Флагнати разговори", "% флагнати"] + [c["label"] for c in CRITERIA]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    detail_sheet_name = "Разговори"
    agent_col = "B"
    category_col = "G"
    status_col = "I"
    overall_col = "J"
    criteria_start_col_idx = 11

    for row_i, agent in enumerate(agent_names, start=2):
        rng_agent = f"'{detail_sheet_name}'!${agent_col}$2:${agent_col}${detail_last_row}"
        rng_status = f"'{detail_sheet_name}'!${status_col}$2:${status_col}${detail_last_row}"
        rng_overall = f"'{detail_sheet_name}'!${overall_col}$2:${overall_col}${detail_last_row}"
        rng_category = f"'{detail_sheet_name}'!${category_col}$2:${category_col}${detail_last_row}"

        agent_cell = f"A{row_i}"
        ws.cell(row=row_i, column=1, value=agent)

        ws.cell(row=row_i, column=2,
                 value=f'=COUNTIFS({rng_agent},{agent_cell},{rng_category},"<>Нерелевантен")')
        ws.cell(row=row_i, column=3,
                 value=f'=ROUND(AVERAGEIFS({rng_overall},{rng_agent},{agent_cell},{rng_category},"<>Нерелевантен"),2)')
        ws.cell(row=row_i, column=4,
                 value=f'=COUNTIFS({rng_agent},{agent_cell},{rng_status},"Флагнат")')
        ws.cell(row=row_i, column=5,
                 value=f'=IFERROR(ROUND(D{row_i}/B{row_i}*100,1),0)')

        for j, c in enumerate(CRITERIA):
            col_letter = get_column_letter(criteria_start_col_idx + j)
            rng_crit = f"'{detail_sheet_name}'!${col_letter}$2:${col_letter}${detail_last_row}"
            out_col = 6 + j
            ws.cell(row=row_i, column=out_col,
                     value=f'=IFERROR(ROUND(AVERAGEIFS({rng_crit},{rng_agent},{agent_cell},{rng_category},"<>Нерелевантен"),2),"-")')

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_i, column=col).font = Font(name=FONT_NAME)

    last_row = ws.max_row
    last_col_letter = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"
    ws.freeze_panes = "A2"

    data_range = f"C2:C{last_row}"
    ws.conditional_formatting.add(
        data_range, FormulaRule(formula=["C2<5"], fill=FLAGGED_FILL)
    )
    ws.conditional_formatting.add(
        data_range, FormulaRule(formula=["C2>=8"], fill=GOOD_FILL)
    )

    ws.column_dimensions["A"].width = 22
    for col_idx in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16


def generate_report():
    analyzed_calls = db.get_all_analyzed_calls()
    if not analyzed_calls:
        print("[export_excel] Няма анализирани разговори в базата — нищо за export.")
        return None

    wb = Workbook()
    detail_ws = wb.active
    detail_ws.title = "Разговори"
    last_row, last_col_letter = build_detail_sheet(detail_ws, analyzed_calls)

    agent_names = sorted({
        c["agent_name"] for c in analyzed_calls
        if c["agent_name"]
        and c["call_category"] != "нерелевантен"
        and c["agent_name"] != "Неизвестен служител"
    })
    summary_ws = wb.create_sheet("Служители - обобщение")
    build_summary_sheet(summary_ws, agent_names, last_row, last_col_letter)

    wb.move_sheet("Служители - обобщение", offset=-1)
    wb.active = 0

    wb.save(OUTPUT_PATH)
    print(f"[export_excel] Записан {OUTPUT_PATH}")

    _recalculate(OUTPUT_PATH)
    return OUTPUT_PATH


def _recalculate(path: Path):
    recalc_script = Path("/mnt/skills/public/xlsx/scripts/recalc.py")
    if not recalc_script.exists():
        print("[export_excel] recalc.py не е наличен в тази среда — "
              "пропускам преизчисление (нормално извън Claude-средата).")
        return
    result = subprocess.run(
        ["python3", str(recalc_script), str(path)],
        capture_output=True, text=True,
    )
    print(f"[export_excel] recalc резултат: {result.stdout.strip()}")
    if result.returncode != 0:
        print(f"[export_excel] ГРЕШКА при recalc: {result.stderr}")


if __name__ == "__main__":
    generate_report()
